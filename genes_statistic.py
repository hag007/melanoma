import Bio.UniProt.GOA as GOA
# from orangecontrib.bio.go import Ontology
import wget
from utils.ensembl2gene_symbol import e2g_convertor
import time
import requests
import scipy.special
import matplotlib.pyplot as plt
from matplotlib import style
import matplotlib.ticker as ticker
style.use("ggplot")
import scipy
from lifelines.statistics import logrank_test
from scipy.stats import hypergeom
from statsmodels.sandbox.stats.multicomp import fdrcorrection0
import logging
sh = logging.StreamHandler()
logger = logging.getLogger("log")
logger.addHandler(sh)
from constants import *
from infra import *
from scipy.stats import pearsonr
from sklearn.cluster import KMeans
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
import sys
import os
from goatools.obo_parser import GODag
from goatools.go_enrichment import GOEnrichmentStudy
from goatools.associations import read_ncbi_gene2go
# from goatools.associations import read_gaf
from lifelines import KaplanMeierFitter
from matplotlib.colors import LinearSegmentedColormap
from scipy.stats import rankdata

############################ () cluster and enrichment #############################


# () main
def genes_statistics(tested_gene_list_file_name, total_gene_list_file_name, gene_expression_file_name, phenotype_file_name, survival_file_name, gene_filter_file_name=None, tested_gene_list_path=None, total_gene_list_path=None, gene_expression_path=None, phenotype_path=None, gene_filter_file_path=None, var_th_index=None, is_unsupervised=True, start_k=2, end_k=2, meta_groups=None, filter_expression=None):

    data = load_integrated_ge_data(tested_gene_list_file_name=tested_gene_list_file_name, total_gene_list_file_name=total_gene_list_file_name, gene_expression_file_name=gene_expression_file_name,                                                                                                                                                    phenotype_file_name=phenotype_file_name, survival_file_name=survival_file_name, var_th_index=var_th_index, meta_groups=meta_groups, filter_expression=filter_expression)
    if data is None:
        print "insufficient data"
        return
    gene_expression_top_var, gene_expression_top_var_headers_rows, gene_expression_top_var_headers_columns, labels_assignment, survival_dataset = data

    plot_genes_statistic(gene_expression_top_var, gene_expression_top_var_headers_columns, tested_gene_list_file_name)



def plot_genes_statistic(gene_expression_top_var, gene_expression_top_var_headers_columns, tested_gene_list_file_name):
    ax = plt.subplot(111)
    positions = np.arange(len(gene_expression_top_var_headers_columns)) + 1
    bp = ax.boxplot(gene_expression_top_var, positions=positions, showmeans=True,
                    labels=e2g_convertor(gene_expression_top_var_headers_columns))
    ax.set_title("genes_statistic_{}_{}_averaged var:{}".format(constants.CANCER_TYPE, tested_gene_list_file_name.split(".")[0], '%.3f' % np.average(np.var(gene_expression_top_var, axis=0))))
    for label in ax.xaxis.get_ticklabels():
        label.set_fontsize(7)
        label.set_rotation(90)
    plt.savefig(os.path.join(constants.BASE_PROFILE, "output",
                             "genes_statistic_{}_{}_{}.png".format(constants.CANCER_TYPE,
                                                                   tested_gene_list_file_name.split(".")[0],
                                                                   time.time())))


def labels_assignments(meta_groups, phenotype_file_name, patients_list):
    labels_assignment = []
    for i, cur_groups in enumerate(meta_groups):
        labeled_patients = divided_patient_ids_by_label(phenotype_file_name, groups=cur_groups)
        cur_labeled = np.array(patients_list)
        for j, cur_patients_group in enumerate(labeled_patients):
            cur_labeled[np.in1d(cur_labeled, cur_patients_group)] = j + 1
        cur_labeled[~np.core.defchararray.isdigit(cur_labeled)] = 0
        labels_assignment.append(cur_labeled.astype(np.int32))
    return labels_assignment



def separate_headers(total_gene_expression):
    total_gene_expression = np.array(total_gene_expression)
    total_gene_expression_headers_columns = total_gene_expression[0][1:]
    total_gene_expression_headers_rows = total_gene_expression[1:, 0]
    total_gene_expression = total_gene_expression[1:]
    total_gene_expression = total_gene_expression[:, 1:]
    total_gene_expression = total_gene_expression.astype(np.float32)
    return total_gene_expression_headers_rows, total_gene_expression_headers_columns, total_gene_expression


def check_enrichment(gene_list):
    ensembl_for_url = re.sub("\.\d{1,2},", ",", gene_list)
    url = "http://david.abcc.ncifcrf.gov/api.jsp?type=ENSEMBL_GENE_ID&ids={}&tool=chartReport&annot=GOTERM_BP_DIRECT,GOTERM_CC_DIRECT,GOTERM_MF_DIRECT,KEGG_PATHWAY".format(ensembl_for_url)
    return url

def print_to_excel(output_rows,gene_list_file_name,gene_expression_file_name,var_th_index):
    wb = Workbook()#ffff00
    ws = wb.active
    yellowFill = PatternFill(start_color='00FFFF00',
                          end_color='00FFFF00',
                          fill_type='solid')
    bd_regular = Side(style='thin', color="000000")
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)

    bd_bold = Side(style='thick', color="000000")
    border_bold = Border(left=bd_bold, top=bd_bold, right=bd_bold, bottom=bd_bold)

    blueDarkFill = PatternFill(start_color='006699FF',
                             end_color='006699FF',
                             fill_type='solid')
    blueMediumFill = PatternFill(start_color='0099CCFF',
                               end_color='0099CCFF',
                               fill_type='solid')
    blueLightFill = PatternFill(start_color='00E6F3FF',
                                 end_color='00E6F3FF',
                                 fill_type='solid')
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)


    headers = ["Description", "Genes", "URL", "GO_NS", "GO_ID", "GO_name", "nominal_pval", "FDR"]
    for i, header in enumerate(headers):
        ws['{}1'.format(chr(65+i))].border = border_regular
        ws['{}1'.format(chr(65+i))].fill = yellowFill
        ws['{}1'.format(chr(65+i))] = header
        ws.column_dimensions['{}'.format(chr(65 + i))].width = 30

    for k, cur in enumerate(headers):
        for i, cur in enumerate(output_rows):
            ws['{}{}'.format(chr(65+k), i+2)].border = border_regular
            ws['{}{}'.format(chr(65+k), i+2)].fill = blueLightFill
            ws['{}{}'.format(chr(65+k), i+2)] = cur[k]
            ws['{}{}'.format(chr(65+k), i+2)].alignment = Alignment(wrap_text=True)

    ws.column_dimensions["{}".format(chr(66+k))].width = 30
    ws["{}1".format(chr(66+k))].border = border_bold
    ws["{}1".format(chr(66+k))].fill = blueDarkFill
    ws["{}1".format(chr(66+k))] = "top var {}".format(var_th_index)
    wb.save(os.path.join(constants.OUTPUT_DIR,"CULSTERS_ENRICHMENT-{}-{}-topvar-{}-{}.xlsx".format(gene_list_file_name, gene_expression_file_name, var_th_index, time.time())))
