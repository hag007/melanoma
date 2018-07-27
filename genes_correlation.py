
import scipy.special
import matplotlib.pyplot as plt
from matplotlib import style
style.use("ggplot")
import scipy
from scipy.stats import hypergeom
from statsmodels.sandbox.stats.multicomp import fdrcorrection0
import logging
sh = logging.StreamHandler()
logger = logging.getLogger("log")
logger.addHandler(sh)
from constants import *
from infra import *
from scipy.stats import pearsonr
from utils.ensembl2gene_symbol import e2g_convertor
from utils.ensembl2gene_symbol import g2e_convertor
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
from utils.hg_test import calc_HG_test
############################ genes correlations #############################


# () main
def find_genes_correlations(tested_gene_list_file_names, total_gene_list_file_name, gene_expression_file_names, intersection_gene_file_names, phenotype_file_name = None, filter_expression=None, var_th_index=None, list_mode="ON_THE_FLY"):
    if filter_expression is not None:
        filtered_patients = [y for x in divided_patient_ids_by_label(phenotype_file_name, groups=filter_expression) for y in x]
    print "about ot analyse: {}".format(str(tested_gene_list_file_names)[:20])
    # fetch gene expression by gene_id, divided by tumor type
    gene_sets = []
    expression_sets = []

    if list_mode == "ON_THE_FLY":
        total_gene_list = total_gene_list_file_name
        intersection_gene_sets = intersection_gene_file_names
    else:
        total_gene_list = load_gene_list(total_gene_list_file_name)
        intersection_gene_sets = []
        if intersection_gene_file_names is not None:
            intersection_gene_sets = [np.array([y.split(".")[0] for y in load_gene_list(x)]) if type(x) == str else [y.split(".")[0] for y in x] for x in intersection_gene_file_names]

    all_gene_expressions = [np.array(load_gene_expression_profile_by_genes(x, gene_expression_file_names[i], list_mode = list_mode)) for i, x in enumerate(tested_gene_list_file_names)]
    if filter_expression is None:
        filtered_patients = np.append(all_gene_expressions[1:], all_gene_expressions[1:])
    mutual_patients = np.array([x for x in all_gene_expressions[0][0][1:] if x in all_gene_expressions[1][0][1:] and x in filtered_patients])
    all_gene_expressions[0] = np.c_[all_gene_expressions[0][:,0], all_gene_expressions[0][:,np.in1d(all_gene_expressions[0][0], mutual_patients)]]
    mutual_patients = np.array([x for x in all_gene_expressions[1][0][1:] if x in all_gene_expressions[0][0][1:] and x in filtered_patients])
    all_gene_expressions[1] = np.c_[all_gene_expressions[1][:,0], all_gene_expressions[1][:,np.in1d(all_gene_expressions[1][0], mutual_patients)]]

    dataset_headers_rows, dataset_headers_columns, dataset = separate_headers(all_gene_expressions[0])
    row_var = np.var(dataset, axis=1)
    row_var_sorted = np.sort(row_var)[::-1]
    if var_th_index is None:
        var_th_index = len(row_var_sorted) - 1
    row_var_th = row_var_sorted[var_th_index]
    row_var_masked_indices = np.where(row_var_th > row_var)[0]
    all_gene_expressions[0] = np.delete(all_gene_expressions[0], row_var_masked_indices, axis=0)

    all_gene_expressions_1 = [[y[0], np.array(y[1:]).astype(np.float)] for x in [all_gene_expressions[0]] for y in x[1:]]
    all_gene_expressions_2 = [[y[0], np.array(y[1:]).astype(np.float)] for x in [all_gene_expressions[1]] for y in x[1:]]


    output = []
    header_columns = []
    for i, cur_1 in enumerate(all_gene_expressions_2):
        header_columns.append(e2g_convertor([all_gene_expressions_2[i][0]])[0])
    for i, cur_1 in enumerate(all_gene_expressions_1):
        for j, cur_2 in enumerate(all_gene_expressions_2):
            prsn = pearsonr(cur_1[1], cur_2[1])
            if not math.isnan(pearsonr(cur_1[1], cur_2[1])[0]):
                output.append([e2g_convertor([all_gene_expressions_1[i][0]])[0], prsn[0], prsn[1]])

    if len(output) == 0:
        return ([], ["{}\t({} {} {} {})".format(1.0, 0, 0, 0, 0) for x in intersection_gene_file_names])
    output = np.array(output)
    fdr_results = fdrcorrection0(output[:,2].astype(np.float32), alpha=0.05, method='indep', is_sorted=False)
    output = np.c_[output, fdr_results[1]]
    output = output[output[:,3].astype(np.float64).argsort(),:]

    hg_scores = []
    for cur_set in intersection_gene_sets:
        # hg_score = calc_HG_test(total_gene_list_N=[x[0].split(".")[0] for x in all_gene_expressions_1], tests_gene_list_B=cur_set, total_gene_list_n=g2e_convertor(output[np.logical_and(output[:, 3].astype(np.float)  < 0.05, output[:, 1].astype(np.float)  < 0)  , 0]))
        hg_score = calc_HG_test(total_gene_list_N=[x.split(".")[0] for x in total_gene_list], tests_gene_list_B=cur_set, total_gene_list_n=g2e_convertor(output[np.logical_and(output[:, 3].astype(np.float)  < 0.05, output[:, 1].astype(np.float)  < 0)  , 0]))
        print hg_score
        hg_scores.append(hg_score)
    file_names = ""
    if  tested_gene_list_file_names[0] is str:
        file_names = "_".join([x.split(".")[0] for x in tested_gene_list_file_names])
    print_to_excel(header_columns, output, intersection_gene_sets, intersection_gene_file_names, file_names)
    return (output, hg_scores)
def print_to_excel(header_columns, data, intersection_gene_sets, intersection_gene_file_names, excel_name):
    wb = Workbook()#ffff00
    ws = wb.active
    yellowFill = PatternFill(start_color='00FFFF00',
                          end_color='00FFFF00',
                          fill_type='solid')
    bd_regular = Side(style='thin', color="000000")
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)

    bd_bold = Side(style='thick', color="000000")
    border_bold = Border(left=bd_bold, top=bd_bold, right=bd_bold, bottom=bd_bold)

    blueDarkFill = PatternFill(start_color='006699FF',
                             end_color='006699FF',
                             fill_type='solid')
    blueMediumFill = PatternFill(start_color='0099CCFF',
                               end_color='0099CCFF',
                               fill_type='solid')
    blueLightFill = PatternFill(start_color='00E6F3FF',
                                 end_color='00E6F3FF',
                                 fill_type='solid')
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)


    for c, cur_gene_set in enumerate([data[:,0]] + intersection_gene_sets):
        filtered_data = [x for x in data if x[0] in cur_gene_set]
        shift = c*5
        for i, header in enumerate(header_columns):
            ws['{}1'.format(chr(shift+66 + i*3))].border = border_regular
            ws['{}1'.format(chr(shift+66 + i*3))].fill = yellowFill
            if c!=0:
                header = "{}_{}".format(header, intersection_gene_file_names[c-1].split(".")[0])
            ws['{}1'.format(chr(shift+66 + i*3))] = header
            ws.merge_cells('{}1:{}1'.format(chr(shift+66 + i*2), chr(shift+68 + i*2)))


        headers = ["ID"] + [y for x in header_columns for y in ["Corr", "P-val", "FDR"]]
        for i, header in enumerate(headers):
            ws['{}2'.format(chr(shift+65+i))].border = border_regular
            ws['{}2'.format(chr(shift+65+i))].fill = yellowFill
            ws['{}2'.format(chr(shift+65+i))] = header
            ws.column_dimensions['{}'.format(chr(shift+65 + i))].width = 10

        for k, cur in enumerate(headers):
            for i, cur in enumerate(filtered_data):
                ws['{}{}'.format(chr(shift+65+k), i+3)].border = border_regular
                ws['{}{}'.format(chr(shift+65+k), i+3)].fill = blueLightFill
                ws['{}{}'.format(chr(shift+65+k), i+3)] = cur[k]
                ws['{}{}'.format(chr(shift+65+k), i+3)].alignment = Alignment(wrap_text=True)
                if k!=0:
                    ws['{}{}'.format(chr(shift + 65 + k), i + 3)] = float(cur[k])
                    ws['{}{}'.format(chr(shift+65 + k), i + 3)].number_format = '0.00E+00'
                if k==1:
                    ws['{}{}'.format(chr(shift+65 + k), i + 3)].number_format = '0.000'

    wb.save(os.path.join(constants.OUTPUT_DIR,"GENES_ENRICHMENT-{}-{}.xlsx".format(excel_name, time.time())))
