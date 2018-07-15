from matplotlib import style
style.use("ggplot")
import logging
sh = logging.StreamHandler()
logger = logging.getLogger("log")
logger.addHandler(sh)
from constants import *
from infra import *
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
import os
from genes_statistic import plot_genes_statistic
from utils.clustering import plot_heatmap
from utils.clustering import find_clusters
from utils.km_curve import km_curve
from utils.ensembl2gene_symbol import e2g_convertor
import spm1d
from utils.pca import plot_pca
from utils.pca import plot_pca_by_samples
############################ () cluster and enrichment #############################


# () main
def patient_sets_distribution_differences(tested_gene_list_file_name, total_gene_list_file_name, gene_expression_file_name, phenotype_file_name, survival_file_name, gene_filter_file_name=None, tested_gene_list_path=None, total_gene_list_path=None, gene_expression_path=None, phenotype_path=None, gene_filter_file_path=None, var_th_index=None, is_unsupervised=True, start_k=2, end_k=2, meta_groups=None, filter_expression=None, clustering_algorithm="euclidean", average_patients=True, compute_hotelling=False):

    data = load_integrated_ge_data(tested_gene_list_file_name=tested_gene_list_file_name, total_gene_list_file_name=total_gene_list_file_name, gene_expression_file_name=gene_expression_file_name,                                                                                                                                                    phenotype_file_name=phenotype_file_name, survival_file_name=survival_file_name, var_th_index=var_th_index, meta_groups=meta_groups, filter_expression=filter_expression)
    if data is None:
        print "insufficient data"
        return
    gene_expression_top_var, gene_expression_top_var_headers_rows, gene_expression_top_var_headers_columns, labels_assignment, survival_dataset = data

    for i, cur_groups in enumerate(meta_groups):
        labeled_patients = divided_patient_ids_by_label(phenotype_file_name, groups=cur_groups)

        ordered_gene_expression = gene_expression_top_var[labels_assignment[i].argsort(), :]
        labels_assignment[i].sort()

        heatmap_values =  gene_expression_top_var


        if average_patients:
            avgs = None
            for cur_label_1 in np.unique(labels_assignment[i]):
                avg = np.average(ordered_gene_expression[np.where(labels_assignment[i] == cur_label_1)[0],:], axis=0)
                if avgs is None:
                    avgs = avg.reshape(1,len(avg))
                else:
                    avgs = np.r_[avgs, avg.reshape(1,len(avg))]
            heatmap_values=avgs

        plot_pca(ordered_gene_expression, [labels_assignment[i]], [meta_groups[i]], tested_gene_list_file_name = tested_gene_list_file_name)
        plot_pca_by_samples(ordered_gene_expression, [labels_assignment[i]], [meta_groups[i]], tested_gene_list_file_name = tested_gene_list_file_name)




        plot_heatmap(heatmap_values, e2g_convertor(gene_expression_top_var_headers_columns),
                     [np.unique(labels_assignment[i])],
                     gene_expression_top_var_headers_rows,
                     tested_gene_list_file_name, label_index=i)


        if compute_hotelling:
            with file(os.path.join(constants.BASE_PROFILE,"output", "hotelling_{}_{}_{}.txt".format(constants.CANCER_TYPE, tested_gene_list_file_name.split(".")[0], time.time())),"w+") as f:
                output = "\t"
                for cur_label_1 in np.unique(labels_assignment[i]):
                    output+= "group {}\t".format(cur_label_1)
                output += "\n"
                for cur_label_1 in np.unique(labels_assignment[i]):
                    output += "group {}\t".format(cur_label_1)
                    for cur_label_2 in np.unique(labels_assignment[i]):
                        if cur_label_2 > cur_label_1: continue

                        cur_label_2_start = np.where(labels_assignment[i]==cur_label_2)[0][0]
                        cur_label_1_end = len(labels_assignment[i])
                        if cur_label_1 != labels_assignment[i][-1]:
                            cur_label_1_end = np.where(labels_assignment[i]==cur_label_1+1)[0][0]
                        cur_label_2_end = len(labels_assignment[i])
                        if cur_label_2 != labels_assignment[i][-1]:
                            cur_label_2_end = np.where(labels_assignment[i] == cur_label_2 + 1)[0][0]

                        T2 = spm1d.stats.hotellings2(ordered_gene_expression[cur_label_1_start:cur_label_1_end], ordered_gene_expression[cur_label_2_start:cur_label_2_end])
                        T2i = T2.inference(0.05)
                        output+="{}\t".format(T2i.p)
                        # km_curve(labeled_patients, survival_dataset[1:], gene_expression_top_var_headers_rows, tested_gene_list_file_name.split(".")[0],label_index=i)
                    output+="\n"
                f.write(output)


def check_enrichment(gene_list):
    ensembl_for_url = re.sub("\.\d{1,2},", ",", gene_list)
    url = "http://david.abcc.ncifcrf.gov/api.jsp?type=ENSEMBL_GENE_ID&ids={}&tool=chartReport&annot=GOTERM_BP_DIRECT,GOTERM_CC_DIRECT,GOTERM_MF_DIRECT,KEGG_PATHWAY".format(ensembl_for_url)
    return url

def print_to_excel(output_rows,gene_list_file_name,gene_expression_file_name,var_th_index):
    wb = Workbook()#ffff00
    ws = wb.active
    yellowFill = PatternFill(start_color='00FFFF00',
                          end_color='00FFFF00',
                          fill_type='solid')
    bd_regular = Side(style='thin', color="000000")
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)

    bd_bold = Side(style='thick', color="000000")
    border_bold = Border(left=bd_bold, top=bd_bold, right=bd_bold, bottom=bd_bold)

    blueDarkFill = PatternFill(start_color='006699FF',
                             end_color='006699FF',
                             fill_type='solid')
    blueMediumFill = PatternFill(start_color='0099CCFF',
                               end_color='0099CCFF',
                               fill_type='solid')
    blueLightFill = PatternFill(start_color='00E6F3FF',
                                 end_color='00E6F3FF',
                                 fill_type='solid')
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)


    headers = ["Description", "Genes", "URL", "GO_NS", "GO_ID", "GO_name", "nominal_pval", "FDR"]
    for i, header in enumerate(headers):
        ws['{}1'.format(chr(65+i))].border = border_regular
        ws['{}1'.format(chr(65+i))].fill = yellowFill
        ws['{}1'.format(chr(65+i))] = header
        ws.column_dimensions['{}'.format(chr(65 + i))].width = 30

    for k, cur in enumerate(headers):
        for i, cur in enumerate(output_rows):
            ws['{}{}'.format(chr(65+k), i+2)].border = border_regular
            ws['{}{}'.format(chr(65+k), i+2)].fill = blueLightFill
            ws['{}{}'.format(chr(65+k), i+2)] = cur[k]
            ws['{}{}'.format(chr(65+k), i+2)].alignment = Alignment(wrap_text=True)

    ws.column_dimensions["{}".format(chr(66+k))].width = 30
    ws["{}1".format(chr(66+k))].border = border_bold
    ws["{}1".format(chr(66+k))].fill = blueDarkFill
    ws["{}1".format(chr(66+k))] = "top var {}".format(var_th_index)
    wb.save(os.path.join(constants.OUTPUT_DIR,"CULSTERS_ENRICHMENT-{}-{}-topvar-{}-{}.xlsx".format(gene_list_file_name, gene_expression_file_name, var_th_index, time.time())))
