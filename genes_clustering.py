import Bio.UniProt.GOA as GOA
# from orangecontrib.bio.go import Ontology
import wget
from utils.go import check_enrichment
from utils.ensembl2entrez import ensembl2entrez_convertor
import time
import requests
import scipy.special
import matplotlib.pyplot as plt
from matplotlib import style
style.use("ggplot")
import scipy
from scipy.stats import hypergeom
from statsmodels.sandbox.stats.multicomp import fdrcorrection0
import logging
sh = logging.StreamHandler()
logger = logging.getLogger("log")
logger.addHandler(sh)
from constants import *
from infra import *
from scipy.stats import pearsonr
from sklearn.cluster import KMeans
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
import sys
import os
from goatools.obo_parser import GODag
from goatools.go_enrichment import GOEnrichmentStudy
from goatools.associations import read_ncbi_gene2go
# from goatools.associations import read_gaf
############################ () cluster and enrichment #############################


# () main
def find_clusters_and_go_enrichment(tested_gene_list_file_name, total_gene_list_file_name, gene_expression_file_name, phenotype_file_name, gene_filter_file_name=None, tested_gene_list_path=None, total_gene_list_path=None, gene_expression_path=None, phenotype_path=None, gene_filter_file_path=None, var_th_index=None, start_k=2, end_k=6):
    # fetch gene expression by gene_id, divided by tumor type
    gene_sets = []
    expression_sets = []
    averaged_expression_sets = []
    tested_gene_expression = load_gene_expression_profile_by_genes(tested_gene_list_file_name, gene_expression_file_name, gene_filter_file_name, total_gene_list_path, gene_expression_path, phenotype_path, gene_filter_file_path)
    tested_gene_expression_headers_rows, tested_gene_expression_headers_columns, tested_gene_expression = separate_headers(tested_gene_expression)
    total_gene_list = load_gene_list(total_gene_list_file_name)
    row_var = np.var(tested_gene_expression,axis=1)
    row_var_sorted = np.sort(row_var)[::-1]
    if var_th_index is not None:
        var_th_index = len(row_var_sorted)
    row_var_th = row_var_sorted[var_th_index]
    row_var_masked_indices = np.where(row_var_th >= row_var)[0]
    gene_expression_top_var = np.delete(tested_gene_expression, row_var_masked_indices, axis=0)
    gene_expression_top_var_headers_rows = np.delete(tested_gene_expression_headers_rows, row_var_masked_indices, axis=0)

    clfs_results = {}
    output_rows = []
    if not os.path.exists(os.path.join(constants.TCGA_DATA_DIR, constants.GO_FILE_NAME)):
        wget.download(constants.GO_OBO_URL, os.path.join(constants.TCGA_DATA_DIR, constants.GO_FILE_NAME))
    # if not os.path.exists(os.path.join(constants.TCGA_DATA_DIR, 'goa_human.gaf')):
    #     wget.download(go_obo_url, os.path.join(constants.TCGA_DATA_DIR, 'goa_human.gaf'))
    obo_dag = GODag(os.path.join(constants.TCGA_DATA_DIR, constants.GO_FILE_NAME))

    assoc = read_ncbi_gene2go(os.path.join(constants.TCGA_DATA_DIR, constants.GO_ASSOCIATION_FILE_NAME), no_top=True)
    g = GOEnrichmentStudy([int(cur) for cur in ensembl2entrez_convertor(total_gene_list)],
                          assoc, obo_dag, methods=["bonferroni", "fdr_bh"])
    g_res = g.run_study([int(cur) for cur in ensembl2entrez_convertor(gene_expression_top_var_headers_rows)])
    GO_results = [(cur.NS, cur.GO, cur.goterm.name, cur.p_uncorrected, cur.p_fdr_bh) for cur in g_res if
                  cur.p_fdr_bh <= 0.05]
    print GO_results

    for n_clusters in range(start_k,end_k+1):
        clfs_results[n_clusters] = []
        km_clf = KMeans(n_clusters).fit(gene_expression_top_var)
        for i in range(n_clusters):
            cluster_indices = np.where(km_clf.labels_!=i)[0]
            gene_expression_cluster = np.delete(gene_expression_top_var_headers_rows, cluster_indices, axis=0)
            gene_headers_row_cluster = np.delete(gene_expression_top_var_headers_rows, cluster_indices, axis=0)
            clfs_results[n_clusters].append((gene_headers_row_cluster, gene_headers_row_cluster))
            desc = "k={} clustering cluster {} has {} genes".format(n_clusters, i, len(gene_expression_cluster))
            gene_list = ",".join(gene_headers_row_cluster)
            url = check_enrichment(gene_list)

            g_res = g.run_study([int(cur) for cur in ensembl2entrez_convertor(gene_headers_row_cluster)])
            GO_results = [(cur.NS, cur.GO, cur.goterm.name, cur.p_uncorrected, cur.p_fdr_bh ) for cur in g_res if cur.p_fdr_bh <= 0.05]

            if len(GO_results)>0:
                go_ns, go_terms, go_names, uncorrectd_pvals, FDRs = zip(*GO_results)
            else:
                go_terms = []
                uncorrectd_pvals= []
                FDRs= []
                go_names = []
                go_ns = []
            output_rows.append((desc, "\r\n".join(gene_headers_row_cluster), url, "\r\n".join(go_ns),
                                "\r\n".join(go_terms), "\r\n".join(go_names) , "\r\n".join(map(str, uncorrectd_pvals)), "\r\n".join(map(str, FDRs))))

    print_to_excel(output_rows=output_rows, gene_list_file_name=tested_gene_list_file_name.split(".")[0], gene_expression_file_name=gene_expression_file_name.split(".")[0], var_th_index=var_th_index)


def print_to_excel(output_rows,gene_list_file_name,gene_expression_file_name,var_th_index):
    wb = Workbook()#ffff00
    ws = wb.active
    yellowFill = PatternFill(start_color='00FFFF00',
                          end_color='00FFFF00',
                          fill_type='solid')
    bd_regular = Side(style='thin', color="000000")
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)

    bd_bold = Side(style='thick', color="000000")
    border_bold = Border(left=bd_bold, top=bd_bold, right=bd_bold, bottom=bd_bold)

    blueDarkFill = PatternFill(start_color='006699FF',
                             end_color='006699FF',
                             fill_type='solid')
    blueMediumFill = PatternFill(start_color='0099CCFF',
                               end_color='0099CCFF',
                               fill_type='solid')
    blueLightFill = PatternFill(start_color='00E6F3FF',
                                 end_color='00E6F3FF',
                                 fill_type='solid')
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)


    headers = ["Description", "Genes", "URL", "GO_NS", "GO_ID", "GO_name", "nominal_pval", "FDR"]
    for i, header in enumerate(headers):
        ws['{}1'.format(chr(65+i))].border = border_regular
        ws['{}1'.format(chr(65+i))].fill = yellowFill
        ws['{}1'.format(chr(65+i))] = header
        ws.column_dimensions['{}'.format(chr(65 + i))].width = 30

    for k, cur in enumerate(headers):
        for i, cur in enumerate(output_rows):
            ws['{}{}'.format(chr(65+k), i+2)].border = border_regular
            ws['{}{}'.format(chr(65+k), i+2)].fill = blueLightFill
            ws['{}{}'.format(chr(65+k), i+2)] = cur[k]
            ws['{}{}'.format(chr(65+k), i+2)].alignment = Alignment(wrap_text=True)

    ws.column_dimensions["{}".format(chr(66+k))].width = 30
    ws["{}1".format(chr(66+k))].border = border_bold
    ws["{}1".format(chr(66+k))].fill = blueDarkFill
    ws["{}1".format(chr(66+k))] = "top var {}".format(var_th_index)
    wb.save(os.path.join(constants.OUTPUT_DIR,"CULSTERS_ENRICHMENT-{}-{}-topvar-{}-{}.xlsx".format(gene_list_file_name, gene_expression_file_name, var_th_index, time.time())))
