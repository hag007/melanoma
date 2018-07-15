import re
import gzip
import shutil
import wget
from download_resources import download
from utils.ensembl2entrez import ensembl2entrez_convertor
from matplotlib import style
style.use("ggplot")
import logging
sh = logging.StreamHandler()
logger = logging.getLogger("log")
logger.addHandler(sh)
from infra import *
import os
from goatools.obo_parser import GODag
from goatools.go_enrichment import GOEnrichmentStudy
from goatools.associations import read_ncbi_gene2go
from utils.ensembl2gene_symbol import e2g_convertor
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment

def check_enrichment(gene_list):
    ensembl_for_url = re.sub("\.\d{1,2},", ",", gene_list)
    url = "http://david.abcc.ncifcrf.gov/api.jsp?type=ENSEMBL_GENE_ID&ids={}&tool=chartReport&annot=GOTERM_BP_DIRECT,GOTERM_CC_DIRECT,GOTERM_MF_DIRECT,KEGG_PATHWAY".format(ensembl_for_url)
    return url


def check_group_enrichment(tested_gene_file_name, total_gene_file_name):
    total_gene_list = load_gene_list(total_gene_file_name)
    tested_gene = load_gene_list(tested_gene_file_name)

    if not os.path.exists(os.path.join(constants.GO_DIR, constants.GO_FILE_NAME)):
        download(constants.GO_OBO_URL, constants.GO_DIR)

    obo_dag = GODag(os.path.join(constants.GO_DIR, constants.GO_FILE_NAME))

    if not os.path.exists(os.path.join(constants.GO_DIR, constants.GO_ASSOCIATION_FILE_NAME)):
        download(constants.GO_ASSOCIATION_GENE2GEO_URL, constants.GO_DIR)
        with gzip.open(os.path.join(constants.GO_DIR, os.path.basename(constants.GO_ASSOCIATION_GENE2GEO_URL)), 'rb') as f_in:
            with open(os.path.join(constants.GO_DIR, constants.GO_ASSOCIATION_FILE_NAME),'wb') as f_out:
                shutil.copyfileobj(f_in, f_out)

    assoc = read_ncbi_gene2go(os.path.join(constants.GO_DIR, constants.GO_ASSOCIATION_FILE_NAME), no_top=True)

    g = GOEnrichmentStudy([int(cur) for cur in ensembl2entrez_convertor(total_gene_list)],
                          assoc, obo_dag, methods=["bonferroni", "fdr_bh"])
    g_res = g.run_study([int(cur) for cur in ensembl2entrez_convertor(tested_gene)])

    GO_results = [(cur.NS, cur.GO, cur.goterm.name, cur.p_uncorrected, cur.p_fdr_bh) for cur in g_res if
                  cur.p_fdr_bh <= 0.05]
    if len(GO_results) > 0:
        go_ns, go_terms, go_names, uncorrectd_pvals, FDRs = zip(*GO_results)
    else:
        go_terms = []
        uncorrectd_pvals = []
        FDRs = []
        go_names = []
        go_ns = []
    output_rows = [("\r\n".join(e2g_convertor(tested_gene)),  "\r\n".join(go_ns),
                        "\r\n".join(go_terms), "\r\n".join(go_names), "\r\n".join(map(str, uncorrectd_pvals)),
                        "\r\n".join(map(str, FDRs)))]
    print_to_excel(output_rows, tested_gene_file_name, total_gene_file_name)

def print_to_excel(output_rows, gene_list_file_name, total_gene_file_name):
    wb = Workbook()  # ffff00
    ws = wb.active
    yellowFill = PatternFill(start_color='00FFFF00',
                             end_color='00FFFF00',
                             fill_type='solid')
    bd_regular = Side(style='thin', color="000000")
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)

    bd_bold = Side(style='thick', color="000000")
    border_bold = Border(left=bd_bold, top=bd_bold, right=bd_bold, bottom=bd_bold)

    blueDarkFill = PatternFill(start_color='006699FF',
                               end_color='006699FF',
                               fill_type='solid')
    blueMediumFill = PatternFill(start_color='0099CCFF',
                                 end_color='0099CCFF',
                                 fill_type='solid')
    blueLightFill = PatternFill(start_color='00E6F3FF',
                                end_color='00E6F3FF',
                                fill_type='solid')
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)

    headers = ["Genes", "GO_NS", "GO_ID", "GO_name", "nominal_pval", "FDR"]
    for i, header in enumerate(headers):
        ws['{}1'.format(chr(65 + i))].border = border_regular
        ws['{}1'.format(chr(65 + i))].fill = yellowFill
        ws['{}1'.format(chr(65 + i))] = header
        ws.column_dimensions['{}'.format(chr(65 + i))].width = 30

    for k, cur in enumerate(headers):
        for i, cur in enumerate(output_rows):
            ws['{}{}'.format(chr(65 + k), i + 2)].border = border_regular
            ws['{}{}'.format(chr(65 + k), i + 2)].fill = blueLightFill
            ws['{}{}'.format(chr(65 + k), i + 2)] = cur[k]
            ws['{}{}'.format(chr(65 + k), i + 2)].alignment = Alignment(wrap_text=True)

    ws.column_dimensions["{}".format(chr(66 + k))].width = 30
    ws["{}1".format(chr(66 + k))].border = border_bold
    ws["{}1".format(chr(66 + k))].fill = blueDarkFill
    wb.save(os.path.join(constants.OUTPUT_DIR,
                         "GENE_SET_ENRICHMENT-{}-{}-{}.xlsx".format(gene_list_file_name.split(".")[0], total_gene_file_name.split(".")[0],
                                                                               time.time())))
