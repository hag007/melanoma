from matplotlib import style
style.use("ggplot")
import logging
sh = logging.StreamHandler()
logger = logging.getLogger("log")
logger.addHandler(sh)
from constants import *
from infra import *
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
import os
from genes_statistic import plot_genes_statistic
from utils.clustering import plot_heatmap
from utils.clustering import find_clusters
from utils.km_curve import km_curve
from utils.ensembl2gene_symbol import e2g_convertor
from utils.hg_test import calc_HG_test

############################ () cluster and enrichment #############################


# () main
def find_clusters_and_survival(tested_gene_list_file_name, total_gene_list_file_name, gene_expression_file_name, phenotype_file_name, survival_file_name, gene_filter_file_name=None, tested_gene_list_path=None, total_gene_list_path=None, gene_expression_path=None, phenotype_path=None, gene_filter_file_path=None, var_th_index=None, is_unsupervised=True, start_k=2, end_k=2, meta_groups=None, filter_expression=None, clustering_algorithm="euclidean"):

    data = load_integrated_ge_data(tested_gene_list_file_name=tested_gene_list_file_name, total_gene_list_file_name=total_gene_list_file_name, gene_expression_file_name=gene_expression_file_name,                                                                                                                                                    phenotype_file_name=phenotype_file_name, survival_file_name=survival_file_name, var_th_index=var_th_index, meta_groups=meta_groups, filter_expression=filter_expression)
    if data is None:
        print "insufficient data"
        return
    gene_expression_top_var, gene_expression_top_var_headers_rows, gene_expression_top_var_headers_columns, labels_assignment, survival_dataset = data

    # plot_genes_statistic(gene_expression_top_var, gene_expression_top_var_headers_columns, tested_gene_list_file_name)

    if is_unsupervised:
        clfs_results = find_clusters(end_k, gene_expression_top_var, gene_expression_top_var_headers_rows,
                                    start_k, e2g_convertor(gene_expression_top_var_headers_columns),
                                   tested_gene_list_file_name, labels_assignment, clustering_algorithm=clustering_algorithm)

        for i in range(start_k,end_k+1):
            km_curve(clfs_results[i], survival_dataset[1:], gene_expression_top_var_headers_rows, tested_gene_list_file_name.split(".")[0],i)
            # B1 = ['TCGA-V4-A9E7-01A', 'TCGA-V4-A9E8-01A', 'TCGA-V4-A9EE-01A', 'TCGA-V4-A9EF-01A', 'TCGA-V4-A9EI-01A', 'TCGA-V4-A9EJ-01A', 'TCGA-V4-A9EK-01A', 'TCGA-V4-A9EL-01A', 'TCGA-V4-A9EQ-01A', 'TCGA-V4-A9ET-01A', 'TCGA-V4-A9EX-01A', 'TCGA-V4-A9F0-01A', 'TCGA-V4-A9F3-01A', 'TCGA-V4-A9F7-01A', 'TCGA-V4-A9F8-01A', 'TCGA-VD-A8KG-01A', 'TCGA-VD-A8KJ-01A', 'TCGA-VD-A8KL-01A', 'TCGA-VD-A8KM-01A', 'TCGA-VD-A8KN-01A', 'TCGA-VD-AA8M-01A', 'TCGA-VD-AA8N-01A', 'TCGA-VD-AA8S-01B', 'TCGA-WC-A87Y-01A', 'TCGA-WC-A880-01A', 'TCGA-WC-A883-01A', 'TCGA-WC-A884-01A', 'TCGA-WC-A885-01A', 'TCGA-WC-A888-01A', 'TCGA-YZ-A980-01A', 'TCGA-YZ-A982-01A', 'TCGA-YZ-A983-01A']
            # B2 = ['TCGA-V4-A9E5-01A', 'TCGA-V4-A9E9-01A', 'TCGA-V4-A9EA-01A', 'TCGA-V4-A9EC-01A', 'TCGA-V4-A9ED-01A', 'TCGA-V4-A9EH-01A', 'TCGA-V4-A9EM-01A', 'TCGA-V4-A9EO-01A', 'TCGA-V4-A9ES-01A', 'TCGA-V4-A9EW-01A', 'TCGA-V4-A9EY-01A', 'TCGA-V4-A9EZ-01A', 'TCGA-V4-A9F1-01A', 'TCGA-V4-A9F2-01A', 'TCGA-V4-A9F4-01A', 'TCGA-VD-A8K7-01B', 'TCGA-VD-A8K9-01A', 'TCGA-VD-A8KA-01B', 'TCGA-VD-A8KB-01A', 'TCGA-VD-A8KE-01A', 'TCGA-VD-A8KH-01A', 'TCGA-VD-A8KK-01A', 'TCGA-VD-A8KO-01A', 'TCGA-VD-AA8P-01A', 'TCGA-VD-AA8R-01A', 'TCGA-VD-AA8T-01A', 'TCGA-WC-A87T-01A', 'TCGA-WC-A87U-01A', 'TCGA-WC-A87W-01A', 'TCGA-WC-A881-01A', 'TCGA-WC-A882-01A', 'TCGA-WC-AA9E-01A', 'TCGA-YZ-A985-01A']
            # N = ['TCGA-V4-A9E7-01A', 'TCGA-V4-A9E8-01A', 'TCGA-V4-A9EE-01A', 'TCGA-V4-A9EF-01A', 'TCGA-V4-A9EI-01A', 'TCGA-V4-A9EJ-01A', 'TCGA-V4-A9EK-01A', 'TCGA-V4-A9EL-01A', 'TCGA-V4-A9EQ-01A', 'TCGA-V4-A9ET-01A', 'TCGA-V4-A9EX-01A', 'TCGA-V4-A9F0-01A', 'TCGA-V4-A9F3-01A', 'TCGA-V4-A9F7-01A', 'TCGA-V4-A9F8-01A', 'TCGA-VD-A8KG-01A', 'TCGA-VD-A8KJ-01A', 'TCGA-VD-A8KL-01A', 'TCGA-VD-A8KM-01A', 'TCGA-VD-A8KN-01A', 'TCGA-VD-AA8M-01A', 'TCGA-VD-AA8N-01A', 'TCGA-VD-AA8S-01B', 'TCGA-WC-A87Y-01A', 'TCGA-WC-A880-01A', 'TCGA-WC-A883-01A', 'TCGA-WC-A884-01A', 'TCGA-WC-A885-01A', 'TCGA-WC-A888-01A', 'TCGA-YZ-A980-01A', 'TCGA-YZ-A982-01A', 'TCGA-YZ-A983-01A', 'TCGA-V4-A9E5-01A', 'TCGA-V4-A9E9-01A', 'TCGA-V4-A9EA-01A', 'TCGA-V4-A9EC-01A', 'TCGA-V4-A9ED-01A', 'TCGA-V4-A9EH-01A', 'TCGA-V4-A9EM-01A', 'TCGA-V4-A9EO-01A', 'TCGA-V4-A9ES-01A', 'TCGA-V4-A9EW-01A', 'TCGA-V4-A9EY-01A', 'TCGA-V4-A9EZ-01A', 'TCGA-V4-A9F1-01A', 'TCGA-V4-A9F2-01A', 'TCGA-V4-A9F4-01A', 'TCGA-VD-A8K7-01B', 'TCGA-VD-A8K9-01A', 'TCGA-VD-A8KA-01B', 'TCGA-VD-A8KB-01A', 'TCGA-VD-A8KE-01A', 'TCGA-VD-A8KH-01A', 'TCGA-VD-A8KK-01A', 'TCGA-VD-A8KO-01A', 'TCGA-VD-AA8P-01A', 'TCGA-VD-AA8R-01A', 'TCGA-VD-AA8T-01A', 'TCGA-WC-A87T-01A', 'TCGA-WC-A87U-01A', 'TCGA-WC-A87W-01A', 'TCGA-WC-A881-01A', 'TCGA-WC-A882-01A', 'TCGA-WC-AA9E-01A', 'TCGA-YZ-A985-01A']
            # print "Group Low HG:"
            # print calc_HG_test(N,B1,clfs_results[i][0])
            # print calc_HG_test(N, B1, clfs_results[i][1])
            # print "Group High HG:"
            # print calc_HG_test(N, B2, clfs_results[i][0])
            # print calc_HG_test(N, B2, clfs_results[i][1])

    else:
        for i, cur_groups in enumerate(meta_groups):
            labeled_patients = divided_patient_ids_by_label(phenotype_file_name, groups=cur_groups)
            plot_heatmap(gene_expression_top_var, e2g_convertor(gene_expression_top_var_headers_columns),
                         [labels_assignment[i]] + labels_assignment[:i] + labels_assignment[i + 1:],
                         gene_expression_top_var_headers_rows,
                         tested_gene_list_file_name, label_index=i)
            km_curve(labeled_patients, survival_dataset[1:], gene_expression_top_var_headers_rows, tested_gene_list_file_name.split(".")[0],label_index=i)



def check_enrichment(gene_list):
    ensembl_for_url = re.sub("\.\d{1,2},", ",", gene_list)
    url = "http://david.abcc.ncifcrf.gov/api.jsp?type=ENSEMBL_GENE_ID&ids={}&tool=chartReport&annot=GOTERM_BP_DIRECT,GOTERM_CC_DIRECT,GOTERM_MF_DIRECT,KEGG_PATHWAY".format(ensembl_for_url)
    return url

def print_to_excel(output_rows,gene_list_file_name,gene_expression_file_name,var_th_index):
    wb = Workbook()#ffff00
    ws = wb.active
    yellowFill = PatternFill(start_color='00FFFF00',
                          end_color='00FFFF00',
                          fill_type='solid')
    bd_regular = Side(style='thin', color="000000")
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)

    bd_bold = Side(style='thick', color="000000")
    border_bold = Border(left=bd_bold, top=bd_bold, right=bd_bold, bottom=bd_bold)

    blueDarkFill = PatternFill(start_color='006699FF',
                             end_color='006699FF',
                             fill_type='solid')
    blueMediumFill = PatternFill(start_color='0099CCFF',
                               end_color='0099CCFF',
                               fill_type='solid')
    blueLightFill = PatternFill(start_color='00E6F3FF',
                                 end_color='00E6F3FF',
                                 fill_type='solid')
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)


    headers = ["Description", "Genes", "URL", "GO_NS", "GO_ID", "GO_name", "nominal_pval", "FDR"]
    for i, header in enumerate(headers):
        ws['{}1'.format(chr(65+i))].border = border_regular
        ws['{}1'.format(chr(65+i))].fill = yellowFill
        ws['{}1'.format(chr(65+i))] = header
        ws.column_dimensions['{}'.format(chr(65 + i))].width = 30

    for k, cur in enumerate(headers):
        for i, cur in enumerate(output_rows):
            ws['{}{}'.format(chr(65+k), i+2)].border = border_regular
            ws['{}{}'.format(chr(65+k), i+2)].fill = blueLightFill
            ws['{}{}'.format(chr(65+k), i+2)] = cur[k]
            ws['{}{}'.format(chr(65+k), i+2)].alignment = Alignment(wrap_text=True)

    ws.column_dimensions["{}".format(chr(66+k))].width = 30
    ws["{}1".format(chr(66+k))].border = border_bold
    ws["{}1".format(chr(66+k))].fill = blueDarkFill
    ws["{}1".format(chr(66+k))] = "top var {}".format(var_th_index)
    wb.save(os.path.join(constants.OUTPUT_DIR,"CULSTERS_ENRICHMENT-{}-{}-topvar-{}-{}.xlsx".format(gene_list_file_name, gene_expression_file_name, var_th_index, time.time())))
