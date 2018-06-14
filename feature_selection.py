from ctypes import c_bool

print(__doc__)
import sys
import constants
import time
from sklearn.svm import SVC
from sklearn.model_selection import StratifiedKFold
from sklearn.feature_selection import RFECV
from sklearn.datasets import make_classification
import fre
from sklearn import datasets, svm
from sklearn.feature_selection import SelectPercentile, f_classif
from scipy.stats import hypergeom
import matplotlib.pyplot as plt
import numpy as np
import os
from sklearn.datasets import load_boston
from sklearn.feature_selection import SelectFromModel
from sklearn.linear_model import LassoCV
from infra import *
import random
import svm as my_svm
import openpyxl
from utils.ensembl2gene_symbol import get_e2g_dictionary
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment

def univariate_feature_selection():
    # #############################################################################
    # Import some data to play with

    # The iris dataset
    iris = datasets.load_iris()

    # Some noisy data not correlated
    E = np.random.uniform(0, 0.1, size=(len(iris.data), 20))

    # Add the noisy data to the informative features
    X = np.hstack((iris.data, E))
    y = iris.target

    plt.figure(1)
    plt.clf()

    X_indices = np.arange(X.shape[-1])

    # #############################################################################
    # Univariate feature selection with F-test for feature scoring
    # We use the default selection function: the 10% most significant features
    selector = SelectPercentile(f_classif, percentile=10)
    selector.fit(X, y)
    scores = -np.log10(selector.pvalues_)
    scores /= scores.max()
    plt.bar(X_indices - .45, scores, width=.2,
            label=r'Univariate score ($-Log(p_{value})$)', color='darkorange',
            edgecolor='black')

    # #############################################################################
    # Compare to the weights of an SVM
    clf = svm.SVC(kernel='linear')
    clf.fit(X, y)

    svm_weights = (clf.coef_ ** 2).sum(axis=0)
    svm_weights /= svm_weights.max()

    plt.bar(X_indices - .25, svm_weights, width=.2, label='SVM weight',
            color='navy', edgecolor='black')

    clf_selected = svm.SVC(kernel='linear')
    clf_selected.fit(selector.transform(X), y)

    svm_weights_selected = (clf_selected.coef_ ** 2).sum(axis=0)
    svm_weights_selected /= svm_weights_selected.max()

    plt.bar(X_indices[selector.get_support()] - .05, svm_weights_selected,
            width=.2, label='SVM weights after selection', color='c',
            edgecolor='black')

    plt.title("Comparing feature selection")
    plt.xlabel('Feature number')
    plt.yticks(())
    plt.axis('tight')
    plt.legend(loc='upper right')
    plt.show()

def load_data(tested_gene_file_names, expression_profile_file_name, phenotype_file_name, label=None, label_values=None, gene_filter_file_name=None):
    genelist_datasets = []
    for tested_gene_file_name in tested_gene_file_names:
        data, labels, _0, gene_ids = my_svm.load_svm_data(tested_gene_file_name, expression_profile_file_name,
                                                 phenotype_file_name, label, label_values,
                                                 gene_filter_file_name)
        genelist_datasets.append(data)

    return (genelist_datasets, labels, gene_ids)

def get_classifier(tuned_parameters, clf=svm.LinearSVC()):
    # GridSearchCV(svm.LinearSVC(probability=True), param_grid=tuned_parameters, return_train_score=True)
    return clf

def rfe(genelist_dataset, labels, gene_list_file_name, tuned_parameters, recursion_step_size=50, start_index = 0, recursion_number_of_steps=20, score_method="average_precision", epoch_iteration=1):
    print "about to start RFE for genelist: {}, recursion step size: {}, score method: {}, epoch_iteration: {}".format(gene_list_file_name.split(".")[0], recursion_step_size, score_method, epoch_iteration)
    classifier = get_classifier(tuned_parameters, svm.LinearSVC())

    rfecv = RFECV(estimator=classifier, step=recursion_step_size, cv=StratifiedKFold(3),
                  scoring=score_method)
    rfecv.fit(genelist_dataset, labels)

    print("Optimal number of features : %d" % rfecv.n_features_)

    # Plot number of features VS. cross-validation scores
    plt.figure()
    plt.xlabel("Number of features selected")
    plt.ylabel("Cross validation score (nb of correct classifications)")
    plt.plot([len(genelist_dataset[0]) % recursion_step_size+ i*recursion_step_size for i in range(len(rfecv.grid_scores_))], rfecv.grid_scores_)
    plt.savefig(os.path.join(constants.OUTPUT_DIR, "rfe_{}_{}_{}_{}.png".format(gene_list_file_name.split(".")[0], recursion_step_size, score_method, epoch_iteration)))
    return (rfecv.n_features_, rfecv.ranking_)

def sparse_elimination_model(genelist_dataset, labels, gene_list_file_name, tuned_parameters, recursion_step_size=50, start_index = 0, recursion_number_of_steps=20, score_method="average_precision", epoch_iteration=1):
    print "about to start RFE for genelist: {}, recursion step size: {}, score method: {}".format(gene_list_file_name.split(".")[0], recursion_step_size, score_method)
    classifier = get_classifier(tuned_parameters, LassoCV())
    classifier.fit(genelist_dataset, labels)
    ranking = [1 if x != 0 else 2 for x in classifier.coef_]
    indices = [i for i, x in enumerate(classifier.coef_)]
    sfm = SelectFromModel(classifier, threshold="mean")
    sfm.fit(genelist_dataset, labels)
    n_features = sfm.transform(genelist_dataset).shape[1]

    # Reset the threshold till the number of features equals two.
    # Note that the attribute can be set directly instead of repeatedly
    # fitting the metatransformer.
    X_transform=genelist_dataset
    while n_features > 1:
        # sfm.threshold += 0.1
        X_transform = sfm.fit_transform(X_transform, labels)
        n_features = X_transform.shape[1]
        indices = [indices[j] for j in sfm.get_support(indices=True)]
        ranking = [x if i in indices else x+1 for i,x in enumerate(ranking)]
        print n_features

    print ranking
    return (n_features, ranking)

    # Plot the selected two features from X.
    # plt.title(
    #     "Features selected from Boston using SelectFromModel with "
    #     "threshold %0.3f." % sfm.threshold)
    # feature1 = X_transform[:, 0]
    # feature2 = X_transform[:, 1]
    # plt.plot(feature1, feature2, 'r.')
    # plt.xlabel("Feature number 1")
    # plt.ylabel("Feature number 2")
    # plt.ylim([np.min(feature2), np.max(feature2)])
    # plt.show()

def tree_feature_selection():
    # Load the boston dataset.
    boston = load_boston()
    X, y = boston['data'], boston['target']

    # We use the base estimator LassoCV since the L1 norm promotes sparsity of features.
    clf = LassoCV()

    # Set a minimum threshold of 0.25
    sfm = SelectFromModel(clf, threshold=0.25)
    sfm.fit(X, y)
    n_features = sfm.transform(X).shape[1]

    # Reset the threshold till the number of features equals two.
    # Note that the attribute can be set directly instead of repeatedly
    # fitting the metatransformer.
    while n_features > 2:
        sfm.threshold += 0.1
        X_transform = sfm.transform(X)
        n_features = X_transform.shape[1]

    # Plot the selected two features from X.
    plt.title(
        "Features selected from Boston using SelectFromModel with "
        "threshold %0.3f." % sfm.threshold)
    feature1 = X_transform[:, 0]
    feature2 = X_transform[:, 1]
    plt.plot(feature1, feature2, 'r.')
    plt.xlabel("Feature number 1")
    plt.ylabel("Feature number 2")
    plt.ylim([np.min(feature2), np.max(feature2)])
    plt.show()

def feature_selection(gene_list_file_name, gene_expression_file_name, phenotype_file_name, groups=None, gene_filter_file_name="protein_coding.txt", rounds=2, recursion_step_size=20, start_index = 0, recursion_number_of_steps=1, feature_selection_method ="rfe", score_method="average_precision", target_genes_subset = None, batches=None, epochs = None):

    genelist_datasets, labels, gene_ids = load_data(gene_list_file_name, gene_expression_file_name, phenotype_file_name, label=None, label_values=None, gene_filter_file_name=None)
    if epochs is None:
        epochs=1
        batches= len(gene_ids)


    thismodule = sys.modules[__name__]
    fsm = getattr(thismodule, feature_selection_method)
    tuned_parameters = {'C': [10], 'kernel': ['linear']}
    genes_subset = None
    if target_genes_subset:
        genes_subset = load_gene_list(target_genes_subset)
    for i, genelist_dataset in enumerate(genelist_datasets):
        for epoch_iteration in range(epochs):
            batch_gene_ids, batch_genelist_dataset = create_batch(genelist_dataset, gene_ids, batches, genes_subset)
            n_features, ranking = fsm(batch_genelist_dataset, labels, gene_list_file_name[i], tuned_parameters, recursion_step_size=recursion_step_size, start_index = 0, recursion_number_of_steps=20, epoch_iteration=epoch_iteration, score_method=score_method)
            sorted_features = sorted(zip(batch_gene_ids, ranking), key = lambda x: x[1])
            features_ranked = zip(*sorted_features)[0]
            print "total ranking: \n{}".format("\n".join(features_ranked))
            if genes_subset:
                genes_subset_ranking = [(j,feature) for feature, j in sorted_features if feature in genes_subset]
                print genes_subset_ranking
                calc_HG_test(genelist_dataset, genes_subset_ranking, ranking)

        print_to_excel(sorted_features=sorted_features, gene_list_file_name=gene_list_file_name[i], feature_selection_method=feature_selection_method, epochs=epochs, batches=batches, recursion_step_size=recursion_step_size)

def calc_HG_test(genelist_dataset, genes_subset_ranking, ranking, th=1):
    b = len(list(filter(lambda x: x[0] == th, genes_subset_ranking)))
    B = len(genes_subset_ranking)
    N = len(genelist_dataset[0])
    n = len(list(filter(lambda x: x == th, ranking)))
    print "run HG test with {},{},{},{}".format(b, N, B, n)
    print hypergeom.sf(b - 1, N, B, n)


def create_batch(genelist_dataset, gene_ids, batches, target_genes_subset):
    genelist_dataset = np.rot90(np.flip(genelist_dataset, 1), k=-1, axes=(1, 0))
    total = zip(gene_ids, genelist_dataset)
    random.shuffle(total)
    gene_ids, genelist_dataset = zip(*total)
    genelist_dataset = list(genelist_dataset)
    gene_ids = list(gene_ids)
    gene_ids_filtered = []
    genelist_dataset_filtered = []
    index=0
    for i, cur in enumerate(list(gene_ids)):
        if cur in target_genes_subset:
            gene_ids_filtered.append(cur)
            genelist_dataset_filtered.append(genelist_dataset[i-index])
            del genelist_dataset[i-index]
            del gene_ids[i-index]
            index+=1
    gene_ids_filtered = gene_ids_filtered + gene_ids[:batches-index]
    genelist_dataset_filtered = genelist_dataset_filtered + genelist_dataset[:batches-index]
    genelist_dataset_filtered = np.flip(np.rot90(genelist_dataset_filtered, k=1, axes=(1, 0)), 1)
    return gene_ids_filtered, genelist_dataset_filtered


def print_to_excel(sorted_features, gene_list_file_name, feature_selection_method, epochs, batches, recursion_step_size):
    wb = Workbook()  # ffff00
    ws = wb.active
    yellowFill = PatternFill(start_color='00FFFF00',
                             end_color='00FFFF00',
                             fill_type='solid')
    bd_regular = Side(style='thin', color="000000")
    border_regular = Border(left=bd_regular, top=bd_regular, right=bd_regular, bottom=bd_regular)

    bd_bold = Side(style='thick', color="000000")
    border_bold = Border(left=bd_bold, top=bd_bold, right=bd_bold, bottom=bd_bold)

    blueDarkFill = PatternFill(start_color='006699FF',
                             end_color='006699FF',
                             fill_type='solid')
    blueMediumFill = PatternFill(start_color='0099CCFF',
                               end_color='0099CCFF',
                               fill_type='solid')
    blueLightFill = PatternFill(start_color='00E6F3FF',
                                 end_color='00E6F3FF',
                                 fill_type='solid')
    gene_symbols_dictionary = get_e2g_dictionary()

    ws['A1'].border = border_regular
    ws['A1'].fill = yellowFill
    ws['A1'] = "index"
    ws['B1'].border = border_regular
    ws['B1'].fill = yellowFill
    ws['B1'] = "ensembl id"
    ws.column_dimensions["B"].width = 20
    ws['C1'].border = border_regular
    ws['C1'].fill = yellowFill
    ws['C1'] = "gene symbol"
    ws['D1'].border = border_regular
    ws['D1'].fill = yellowFill
    ws['D1'] = "rank"

    ws['F1'].border = border_bold
    ws['F1'].fill = yellowFill
    ws['F1'] = "{} features chosen out of {}".format(len(list(filter(lambda x: x[1] == 1, sorted_features))), len(sorted_features))

    ws['F2'].border = border_bold
    ws['F2'].fill = yellowFill
    ws['F2'] = "step size = {}".format(recursion_step_size)

    ws.column_dimensions["F"].width = 30

    for i, feature in enumerate(sorted_features):

        ws['A{}'.format(i+2)].border = border_regular
        ws['A{}'.format(i+2)].fill = blueLightFill
        ws['A{}'.format(i+2)] = str(i+1)
        ws['A{}'.format(i+2)].alignment = Alignment(horizontal='center', wrap_text=True)

        ws['B{}'.format(i+2)].border = border_regular
        ws['B{}'.format(i+2)].fill = blueLightFill
        ws['B{}'.format(i+2)] = feature[0]
        ws['B{}'.format(i+2)].alignment = Alignment(horizontal='center', wrap_text=True)

        ws['C{}'.format(i + 2)].border = border_regular
        ws['C{}'.format(i + 2)].fill = blueLightFill
        ws['C{}'.format(i + 2)] = gene_symbols_dictionary[feature[0].split(".")[0]]
        ws['C{}'.format(i + 2)].alignment = Alignment(horizontal='center', wrap_text=True)

        ws['D{}'.format(i+2)].border = border_regular
        ws['D{}'.format(i+2)].fill = blueLightFill
        ws['D{}'.format(i+2)] = feature[1]
        ws['D{}'.format(i+2)].alignment = Alignment(horizontal='center', wrap_text=True)

    wb.save(os.path.join(constants.OUTPUT_DIR,"FEATURE-SELECTION-{}-{}-{}-{}-{}.xlsx".format(gene_list_file_name.split(".")[0], feature_selection_method, epochs, batches, time.time())))